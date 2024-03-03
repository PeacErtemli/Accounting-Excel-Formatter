from openpyxl import load_workbook


def readExcel():
    wb = load_workbook("C:/Users/barisertemli/Desktop/annemExcelAlgoritma/Diger_Ekim.xlsx")
    ws = wb.active
    header_range = ws["A1:AA1"][0]
    # header_range = ws.rows
    # TODO: Range'te [0] bulunmasının sebebi sheet'ten 2 indexli bir liste dönmesi. İlki verilerle dolu, ikincisi boş. Neden olduğunu araştır.
    # TODO: Find a better way to iterate through rows.

    # print(header_list)
    return header_range


def menu(header_range):
    header_list = list(header_range)

    done = False
    menuDict = {}
    counter = 0
    for cell in header_list:
        counter += 1
        menuDict[counter] = cell.value

    while not done:
        print("\n")
        print("Welcome to the Excel Formatter!")

        for key, value in menuDict.items():
            print(f"{key} - {value}")

        try:
            delete_input = int(input("Press the header number that you want to DELETE: \n Press 0 to exit"))
            # TODO: Figure out to split if the input is integer ot string
            if delete_input == 0:
                done = True
            else:
                menuDict.pop(delete_input)
        except ValueError:
            print("Please provide an integer.")
